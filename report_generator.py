"""
Report Generator - Crea Excel e HTML per email
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any
from logger_config import setup_logger

logger = setup_logger('ReportGenerator')


class ReportGenerator:
    """Genera report Excel e HTML"""

    def __init__(self, output_dir: str = "reports"):
        """
        Inizializza generator

        Args:
            output_dir: Directory output
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.logger = logger

    def generate_excel(self, data: Dict[str, Any]) -> str:
        """
        Genera report Excel di base (per compatibilità)

        Args:
            data: Dati completi report

        Returns:
            str: Path file Excel
        """
        try:
            # Crea workbook
            wb = openpyxl.Workbook()

            # Sheet 1: Summary
            self._create_summary_sheet(wb, data)

            # Sheet 2: Top Defects
            self._create_defects_sheet(wb, data)

            # Sheet 3: AI Insights
            self._create_insights_sheet(wb, data)

            # Salva
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"AI_Scrap_Report_{timestamp}.xlsx"
            filepath = self.output_dir / filename

            wb.save(filepath)
            self.logger.info(f"✅ Excel generato: {filepath}")

            return str(filepath)

        except Exception as e:
            self.logger.error(f"❌ Errore generazione Excel: {e}")
            raise

    def _create_summary_sheet(self, wb, data):
        """Crea sheet Summary"""
        ws = wb.active
        ws.title = "Summary"

        # Header
        ws['A1'] = "AI SCRAP ANALYSIS REPORT"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells('A1:D1')

        # Info
        row = 3
        ws[f'A{row}'] = "Periodo:"
        ws[f'B{row}'] = data.get('period', 'N/A')
        ws[f'B{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Data Report:"
        ws[f'B{row}'] = datetime.now().strftime('%d/%m/%Y %H:%M')

        # Statistiche
        row += 2
        prod = data.get('production_data', {})

        stats = [
            ("Ordini Prodotti", prod.get('NrOrders', 0)),
            ("Schede Prodotte", prod.get('NrBoards', 0)),
            ("Difetti Totali", data.get('scrap_count', 0)),
            ("Scrap Rate %", data.get('analysis', {}).get('scrap_rate', 0))
        ]

        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            if isinstance(value, float):
                ws[f'B{row}'].number_format = '0.00'
            row += 1

        # Auto-width
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

    def _create_defects_sheet(self, wb, data):
        """Crea sheet Top Defects"""
        ws = wb.create_sheet("Top Defects")

        # Header
        headers = ["Rank", "Defect Name", "Count", "Percentage %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Dati
        defects = data.get('top_defects', [])
        total = sum(d['Count'] for d in defects)

        for idx, defect in enumerate(defects[:20], 2):
            ws.cell(idx, 1, idx - 1)
            ws.cell(idx, 2, defect.get('DefectName', 'N/A'))
            ws.cell(idx, 3, defect.get('Count', 0))

            percentage = (defect['Count'] / total * 100) if total > 0 else 0
            cell = ws.cell(idx, 4, percentage)
            cell.number_format = '0.00'

        # Chart
        chart = BarChart()
        chart.title = "Top 10 Defects"
        chart.x_axis.title = "Defect"
        chart.y_axis.title = "Count"

        data_ref = Reference(ws, min_col=3, min_row=1, max_row=11)
        cats_ref = Reference(ws, min_col=2, min_row=2, max_row=11)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        ws.add_chart(chart, "F2")

        # Auto-width
        for col in range(1, 5):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    def _create_insights_sheet(self, wb, data):
        """Crea sheet AI Insights"""
        ws = wb.create_sheet("AI Insights")

        # Header
        ws['A1'] = "ROOT CAUSE ANALYSIS"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        ws.merge_cells('A1:C1')

        # Root Causes
        row = 3
        insights = data.get('ai_insights', {})
        root_causes = insights.get('root_causes', [])

        for cause in root_causes:
            ws[f'A{row}'] = cause.get('category', 'N/A')
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

            ws[f'B{row}'] = cause.get('cause', 'N/A')
            ws[f'C{row}'] = cause.get('impact', 'N/A')
            row += 1

        # Recommendations
        row += 2
        ws[f'A{row}'] = "RECOMMENDATIONS"
        ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws.merge_cells(f'A{row}:C{row}')

        row += 2
        recommendations = insights.get('recommendations', [])

        for rec in recommendations:
            ws[f'A{row}'] = rec.get('title', 'N/A')
            ws[f'A{row}'].font = Font(bold=True)

            ws[f'B{row}'] = rec.get('description', 'N/A')
            ws[f'C{row}'] = rec.get('priority', 'Media')

            # Colore priorità
            priority = rec.get('priority', 'Media')
            if priority == 'Alta' or priority == 'Critica':
                ws[f'C{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws[f'C{row}'].font = Font(color="FFFFFF", bold=True)
            elif priority == 'Media':
                ws[f'C{row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

            row += 1

        # Auto-width
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15

    def generate_html_body(self, data: Dict[str, Any]) -> str:
        """
        Genera corpo HTML per email

        Args:
            data: Dati report

        Returns:
            str: HTML formattato
        """
        # Estrai dati
        period = data.get('period', 'N/A')
        production = data.get('production_data', {})
        analysis = data.get('analysis', {})
        ai_insights = data.get('ai_insights', {})
        top_defects = data.get('top_defects', [])

        # Statistiche
        nr_orders = production.get('NrOrders', 0)
        nr_boards = production.get('NrBoards', 0)
        scrap_count = data.get('scrap_count', 0)
        scrap_rate = analysis.get('scrap_rate', 0)

        # Colore scrap rate
        if scrap_rate < 2.0:
            scrap_color = '#28a745'
        elif scrap_rate < 5.0:
            scrap_color = '#ffc107'
        else:
            scrap_color = '#dc3545'

        # Root causes e recommendations
        root_causes = ai_insights.get('root_causes', [])
        recommendations = ai_insights.get('recommendations', [])

        # Costruisci HTML
        html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 900px;
            margin: 0 auto;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            background-color: white;
            border-radius: 8px;
            padding: 30px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            border-radius: 8px 8px 0 0;
            margin: -30px -30px 30px -30px;
            text-align: center;
        }}
        .header h1 {{
            margin: 0;
            font-size: 28px;
            font-weight: 600;
        }}
        .header p {{
            margin: 10px 0 0 0;
            font-size: 16px;
            opacity: 0.9;
        }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 20px;
            margin: 30px 0;
        }}
        .stat-card {{
            background: #f8f9fa;
            padding: 20px;
            border-radius: 8px;
            border-left: 4px solid #667eea;
        }}
        .stat-card h3 {{
            margin: 0 0 10px 0;
            color: #666;
            font-size: 14px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        .stat-card .value {{
            font-size: 32px;
            font-weight: bold;
            color: #333;
        }}
        .scrap-rate {{
            background: {scrap_color};
            color: white;
            padding: 15px;
            border-radius: 8px;
            text-align: center;
            margin: 20px 0;
        }}
        .scrap-rate h3 {{
            margin: 0 0 10px 0;
            font-size: 16px;
            opacity: 0.9;
        }}
        .scrap-rate .value {{
            font-size: 48px;
            font-weight: bold;
            margin: 0;
        }}
        .section {{
            margin: 30px 0;
        }}
        .section h2 {{
            color: #667eea;
            border-bottom: 2px solid #667eea;
            padding-bottom: 10px;
            margin-bottom: 20px;
            font-size: 22px;
        }}
        .defect-list {{
            background: #f8f9fa;
            padding: 20px;
            border-radius: 8px;
        }}
        .defect-item {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 12px;
            margin: 8px 0;
            background: white;
            border-radius: 6px;
            border-left: 4px solid #dc3545;
        }}
        .defect-name {{
            font-weight: 600;
            color: #333;
        }}
        .defect-count {{
            background: #dc3545;
            color: white;
            padding: 4px 12px;
            border-radius: 20px;
            font-weight: bold;
        }}
        .insight-box {{
            background: #e7f3ff;
            border-left: 4px solid #2196F3;
            padding: 15px;
            margin: 15px 0;
            border-radius: 4px;
        }}
        .insight-box h4 {{
            margin: 0 0 10px 0;
            color: #1976D2;
            font-size: 16px;
        }}
        .recommendation-box {{
            background: #f0f9ff;
            border-left: 4px solid #28a745;
            padding: 15px;
            margin: 15px 0;
            border-radius: 4px;
        }}
        .recommendation-box h4 {{
            margin: 0 0 10px 0;
            color: #28a745;
            font-size: 16px;
        }}
        .footer {{
            margin-top: 40px;
            padding-top: 20px;
            border-top: 2px solid #eee;
            text-align: center;
            color: #666;
            font-size: 14px;
        }}
        .attachments {{
            background: #fff3cd;
            border: 1px solid #ffc107;
            padding: 15px;
            border-radius: 8px;
            margin: 20px 0;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 AI Scrap Analysis Report</h1>
            <p>VANDEWIELE ROMANIA SRL - PTHM Department</p>
            <p>Periodo: {period}</p>
        </div>

        <div class="stats-grid">
            <div class="stat-card">
                <h3>📦 Ordini Prodotti</h3>
                <div class="value">{nr_orders}</div>
            </div>
            <div class="stat-card">
                <h3>🔧 Schede Prodotte</h3>
                <div class="value">{nr_boards:,}</div>
            </div>
        </div>

        <div class="scrap-rate">
            <h3>Scrap Rate</h3>
            <div class="value">{scrap_rate:.2f}%</div>
            <p style="margin: 10px 0 0 0; opacity: 0.9;">{scrap_count} difetti totali</p>
        </div>

        <div class="section">
            <h2>🔝 Top 10 Difetti</h2>
            <div class="defect-list">
"""

        # Top defects
        for defect in top_defects[:10]:
            defect_name = defect.get('DefectName', 'N/A')
            count = defect.get('Count', 0)
            html += f"""
                <div class="defect-item">
                    <span class="defect-name">{defect_name}</span>
                    <span class="defect-count">{count}</span>
                </div>
"""

        html += """
            </div>
        </div>
"""

        # Root Causes
        if root_causes:
            html += """
        <div class="section">
            <h2>🔍 Root Cause Analysis (AI)</h2>
"""
            for cause in root_causes[:5]:
                html += f"""
            <div class="insight-box">
                <h4>🎯 {cause.get('category', 'Analisi')}</h4>
                <p><strong>Causa:</strong> {cause.get('cause', 'N/A')}</p>
                <p><strong>Impatto:</strong> {cause.get('impact', 'N/A')}</p>
            </div>
"""
            html += """
        </div>
"""

        # Recommendations
        if recommendations:
            html += """
        <div class="section">
            <h2>💡 Raccomandazioni (AI)</h2>
"""
            for rec in recommendations[:5]:
                html += f"""
            <div class="recommendation-box">
                <h4>✅ {rec.get('title', 'Raccomandazione')}</h4>
                <p>{rec.get('description', 'N/A')}</p>
                <p><strong>Priorità:</strong> {rec.get('priority', 'Media')}</p>
            </div>
"""
            html += """
        </div>
"""

        # Footer
        html += f"""
        <div class="attachments">
            <h4>📎 File Allegati</h4>
            <ul>
                <li>📊 Report Excel completo con analisi dettagliata</li>
            </ul>
        </div>

        <div class="footer">
            <p><strong>VANDEWIELE ROMANIA SRL</strong></p>
            <p>Report generato automaticamente il {datetime.now().strftime('%d/%m/%Y alle %H:%M:%S')}</p>
        </div>
    </div>
</body>
</html>
"""

        return html
