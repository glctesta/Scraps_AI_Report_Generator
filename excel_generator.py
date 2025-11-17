"""
Module for generating advanced Excel reports with charts and summaries.
Uses openpyxl to create professionally formatted Excel files.
"""
import pandas as pd
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from openpyxl.cell import MergedCell
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from logger_config import setup_logger

logger = setup_logger('ExcelGenerator')


class ExcelReportGenerator:
    """Class to generate styled Excel reports with charts."""

    def __init__(self, title="AI Analysis Report"):
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl is not available. Please install with: pip install openpyxl pandas")
            raise ImportError("openpyxl or pandas not installed")

        self.title_text = title
        self._setup_styles()
        logger.info("ExcelReportGenerator initialized.")

    def _setup_styles(self):
        """Defines styles for the Excel document."""
        self.header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        self.title_font = Font(name='Calibri', size=18, bold=True, color='1F4E78')
        self.subtitle_font = Font(name='Calibri', size=14, bold=True, color='44546A')
        self.chart_title_font = Font(name='Calibri', size=12, bold=True)

    def generate_report(self, report_data: dict, output_path: str) -> str:
        """
        Generates a complete Excel report from standardized analysis data, including charts.
        """
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # --- Create Sheets ---
            self._create_summary_sheet(wb, report_data)
            self._create_charts_sheet(wb, report_data) # New sheet for charts

            if report_data.get('ytd_data'):
                self._create_ytd_sheet(wb, report_data) # New sheet for Year-to-Date data

            if report_data.get('root_causes'):
                self._create_dataframe_sheet(wb, "AI Root Causes", pd.DataFrame(report_data['root_causes']))
            if report_data.get('recommendations'):
                self._create_dataframe_sheet(wb, "AI Recommendations", pd.DataFrame(report_data['recommendations']))
            if report_data.get('raw_data'):
                self._create_dataframe_sheet(wb, "Raw Data", pd.DataFrame(report_data['raw_data']))

            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            logger.info(f"Excel report saved successfully to: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error during Excel report generation: {e}", exc_info=True)
            return ""

    def _auto_fit_columns(self, ws, min_width=12, max_width=50):
        """Adjusts column widths based on content, safely ignoring merged cells."""
        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[column_letter]:
                if isinstance(cell, MergedCell): continue
                if cell.value:
                    try:
                        cell_length = max(len(line) for line in str(cell.value).split('\n'))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
            adjusted_width = max(min_width, max_length + 4)
            final_width = min(adjusted_width, max_width)
            ws.column_dimensions[column_letter].width = final_width

    def _create_summary_sheet(self, wb: Workbook, data: dict):
        """Creates the main summary sheet with key metrics."""
        ws = wb.create_sheet("Summary", 0)
        ws['A1'] = data.get('analysis_type', "Analysis Report")
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')

        ws['A3'] = "Period"
        ws['B3'] = data.get('period', 'N/A')
        ws['A4'] = "Generated On"
        ws['B4'] = data.get('generation_date', 'N/A')

        ws['A6'] = "Executive Summary"
        ws['A6'].font = self.subtitle_font
        summary_cell = ws['A7']
        summary_cell.value = data.get('executive_summary', 'Not available.')
        summary_cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells('A7:E15')

        # Key Metrics table
        stats = data.get('statistics', {})
        metrics = []
        if 'scrap_rate' in stats:
            metrics.extend([('Total Scraps', stats.get('total_scraps', 0)), ('Scrap Rate', f"{stats.get('scrap_rate', 0):.2f}%")])
        if 'fail_rate' in stats:
            metrics.extend([('Total Fails', stats.get('total_fails', 0)), ('Fail Rate', f"{stats.get('fail_rate', 0):.2f}%")])
        if 'total_downtime_hours' in stats:
            metrics.extend([('Total Stoppages', stats.get('total_stoppages', 0)), ('Total Downtime', f"{stats.get('total_downtime_hours', 0):.2f} hrs")])

        ws['G3'] = "Key Metrics"
        ws['G3'].font = self.subtitle_font
        for i, (key, value) in enumerate(metrics, 4):
            ws[f'G{i}'] = key
            ws[f'H{i}'] = value

        self._auto_fit_columns(ws)

    def _create_charts_sheet(self, wb: Workbook, data: dict):
        """Creates a new sheet dedicated to charts using a standardized 'chart_data' key."""
        ws = wb.create_sheet("Charts", 1)
        ws['A1'] = f"{data.get('analysis_type', '')} - Visual Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:Q1')

        # --- NUOVA LOGICA ---
        # Cerca la chiave standard 'chart_data'
        chart_data_list = data.get('chart_data', [])

        if chart_data_list:
            # Prepara i dati nel foglio per i grafici
            # Assume che chart_data_list sia una lista di dizionari con chiavi 'label' e 'value'
            ws['A3'] = "Top 5 Issues"
            ws['B3'] = "Count"
            for i, item in enumerate(chart_data_list[:5], 4):
                ws[f'A{i}'] = item.get('label', 'N/A')
                ws[f'B{i}'] = item.get('value', 0)

            # --- Bar Chart ---
            bar_chart = BarChart()
            bar_chart.title = "Top 5 Issues by Frequency"
            bar_chart.style = 11

            chart_data_ref = Reference(ws, min_col=2, min_row=3, max_row=8)
            categories_ref = Reference(ws, min_col=1, min_row=4, max_row=8)
            bar_chart.add_data(chart_data_ref, titles_from_data=True)
            bar_chart.set_categories(categories_ref)
            bar_chart.legend = None
            ws.add_chart(bar_chart, "D3")

            # --- Pie Chart ---
            pie_chart = PieChart()
            pie_chart.title = "Distribution of Top 5 Issues"
            pie_chart.style = 4

            pie_data_ref = Reference(ws, min_col=2, min_row=4, max_row=8)
            labels_ref = Reference(ws, min_col=1, min_row=4, max_row=8)
            pie_chart.add_data(pie_data_ref)
            pie_chart.set_categories(labels_ref)
            ws.add_chart(pie_chart, "L3")
        else:
            ws['A3'] = "No data available for charting."

    def _create_ytd_sheet(self, wb: Workbook, data: dict):
        """Creates the Year-to-Date analysis sheet."""
        ws = wb.create_sheet("Year-to-Date Analysis", 2)
        ws['A1'] = f"{data.get('analysis_type', '')} - Year-to-Date Trend"
        ws['A1'].font = self.title_font

        ytd_data = data.get('ytd_data')
        if ytd_data:
            df = pd.DataFrame(ytd_data)
            # Ensure 'Month' is sorted correctly
            df['MonthNum'] = pd.to_datetime(df['Month'], format='%Y-%m').dt.month
            df = df.sort_values('MonthNum')
            df = df.drop(columns=['MonthNum'])

            # Write data to sheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # Apply styles
            ws['A2'].parent.title = "YTD Data"
            header_cells = ws[1]
            for cell in header_cells:
                cell.font = self.header_font
                cell.fill = self.header_fill

            # --- Line Chart for YTD Trend ---
            line_chart = LineChart()
            line_chart.title = "Monthly Trend (Year-to-Date)"
            line_chart.style = 13
            line_chart.y_axis.title = "Count / Rate"
            line_chart.x_axis.title = "Month"

            # Select columns to plot (all except the 'Month')
            data_cols = [col for col in df.columns if col != 'Month']

            chart_data = Reference(ws, min_col=2, max_col=len(df.columns), min_row=1, max_row=len(df) + 1)
            categories = Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)

            line_chart.add_data(chart_data, titles_from_data=True)
            line_chart.set_categories(categories)

            ws.add_chart(line_chart, "F2")

        else:
            ws['A3'] = "No Year-to-Date data available."

    def _create_dataframe_sheet(self, wb: Workbook, sheet_name: str, df: pd.DataFrame):
        """Creates a new sheet from a pandas DataFrame and styles it."""
        if df.empty:
            return

        ws = wb.create_sheet(sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Apply styles
        ws['A1'].font = self.header_font # The dataframe headers are now the sheet headers
        ws['A1'].fill = self.header_fill
        header_cells = ws[1]
        for cell in header_cells:
            cell.font = self.header_font
            cell.fill = self.header_fill

        self._auto_fit_columns(ws)